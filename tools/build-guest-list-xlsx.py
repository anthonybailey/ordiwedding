#!/usr/bin/env python3
"""
Build a one-time guest-list XLSX for the wedding admin sheet.

PURPOSE
    Generate a single .xlsx file that Anthony imports into Google Sheets
    (File -> Import -> Upload). After import the Sheet is the canonical
    source of truth -- Anthony and Julie edit it manually thereafter.
    This script is dormant unless we choose to regenerate from updated data.

USAGE
    pip install openpyxl
    python3 tools/build-guest-list-xlsx.py
    # Output: tools/guest-list.xlsx

DATA SOURCES (in increasing order of authority for conflicts)
    1. Julie's working CSV ("scratchy", largest superset, ~60 names, lots of
       blanks and inconsistencies; provided in chat 2026-04-07)
    2. Julie's email list of guests not yet on the website / FB event
       (17 parties; provided in chat 2026-04-07)
    3. Form responses on monkeys at /-/rsvp-data/ (21 real + 2 tests;
       fetched via HTTP 2026-04-07)
    Form data wins on conflicts because it's the most direct evidence
    from the guest themselves.

DECISIONS LOG (Anthony confirmations on 2026-04-07)
    - One row per party (NOT per person).
    - Columns: Party, Headcount, Contact, RSVP, Source, Museum (AM),
      Dietary, Accommodation, Travel, Notes, Message
    - Dropdowns on RSVP, Source, Museum (preserved via openpyxl
      DataValidation -- this is why XLSX rather than CSV)
    - Sheet 2 "Links": Facebook event URL + key ordiwedd.ing pages
    - Test form entries (Claude, Anthony+Julie) are KEPT
    - All CSV names are invited (no "not yet invited" category)
    - CSV "FB=Y" means accepted via Facebook -> RSVP=yes, Source=Facebook
    - CSV "FB=M" means maybe via Facebook -> RSVP=maybe, Source=Facebook
    - CSV "ODW=Y" means responded via website -> Source=form
      (cross-checked against actual fetched form data)
    - Reserve list (CSV lines 57-61): Ralph, Rhys, Diane, Tammi Dallaston,
      Terri Wilson -> tagged "Reserve list" in notes, kept in sheet
    - Ash + Alexis are a couple. CSV's "Amber" appears to be an error,
      written as "Alexis" with a flag note for Julie to verify.
    - Best-effort party groupings flagged in notes for Julie review.
    - "Source" left blank where the channel is unknown (don't guess).

PENDING (not yet captured)
    - "Invited via Facebook (no response)" name ranges from Anthony.
      These will become rows with RSVP=awaiting and Source value
      "FB invited (no response)".
    - Julie to verify the tentative groupings flagged in notes.

PROVENANCE
    Each party dict has a `_src` tuple listing source identifiers
    (form filenames, CSV row numbers, "julie-email"). This is for
    traceability and is NOT written to the XLSX.
"""

from pathlib import Path

OUTPUT = Path(__file__).parent / "guest-list.xlsx"

COLUMNS = [
    "Party",
    "Headcount",
    "Contact",
    "RSVP",
    "Source",
    "Museum (AM)",
    "Dietary",
    "Accommodation",
    "Travel",
    "Notes",
    "Message",
]

# Dropdown vocabularies
RSVP_VALUES = ["yes", "no", "maybe", "awaiting"]
SOURCE_VALUES = [
    "form",
    "email",
    "text",
    "verbal",
    "Facebook",
    "FB invited (no response)",
]
MUSEUM_VALUES = ["yes", "no", "maybe"]

# Links tab content
LINKS = [
    ("Facebook event", "https://www.facebook.com/events/1271646081447068/"),
    ("Wedding homepage", "https://ordiwedd.ing/"),
    ("RSVP form", "https://ordiwedd.ing/-/rsvp"),
    ("RSVP responses (raw)", "https://ordiwedd.ing/-/rsvp-responses"),
    ("Our Story", "https://ordiwedd.ing/-/our-story"),
    ("Gifts", "https://ordiwedd.ing/-/gifts"),
    ("Travel", "https://ordiwedd.ing/-/travel"),
]


def party(
    name,
    headcount=1,
    contact="",
    rsvp="awaiting",
    source="",
    museum="",
    dietary="",
    accommodation="",
    travel="",
    notes="",
    message="",
    _src=(),
):
    """Build a party row dict. _src is provenance, not written to XLSX."""
    return dict(
        party=name,
        headcount=headcount,
        contact=contact,
        rsvp=rsvp,
        source=source,
        museum=museum,
        dietary=dietary,
        accommodation=accommodation,
        travel=travel,
        notes=notes,
        message=message,
        _src=list(_src),
    )


# ----------------------------------------------------------------------------
# GUEST DATA
# ----------------------------------------------------------------------------
# Order roughly follows Julie's CSV. Parties from Julie's email that are
# not in the CSV are appended at the end. Form-only test entries (Claude)
# are at the very end.
# ----------------------------------------------------------------------------

PARTIES = [
    # ========================================================================
    # The couple (CSV rows 2-3, also a form test submission)
    # ========================================================================
    party(
        "Anthony Bailey & Julie Dawson",
        headcount=2,
        contact="mail@anthonybailey.net / julie.dawson@gmail.com",
        rsvp="yes",
        source="form",
        museum="yes",
        notes="(it's their wedding)",
        message=(
            "Updated RSVP to test the new museum question. Hi Julie! "
            "This is Anthony and Claude testing the form - you should see "
            "a museum field now."
        ),
        _src=("form:yes-2-anthony_bailey+julie_dawson.txt", "csv:r2-3"),
    ),

    # ========================================================================
    # CSV-order parties
    # ========================================================================

    # CSV r4 -- Alice Dawson, also form respondent
    party(
        "Alice Dawson",
        headcount=1,
        contact="afdawson38@gmail.com",
        rsvp="yes",
        source="form",
        museum="yes",
        message="I hope the day will be all you want it to be---even the weather!",
        _src=("form:yes-1-alice_dawson.txt", "csv:r4"),
    ),

    # CSV r5 -- FB invited, no response
    party(
        "Brandon Ritchey",
        source="FB invited (no response)",
        _src=("csv:r5", "fb-invited:Brandon-Elijah"),
    ),

    # CSV r6 -- possibly = "Elizabeth" form respondent? Tentative match.
    party(
        "Elizabeth Hebden",
        headcount=1,
        contact="07984016674",
        rsvp="yes",
        source="form",
        museum="yes",
        dietary="I eat anything apart from cottage cheese, liver, crab and oysters.",
        notes=(
            "(form 'Elizabeth' tentatively matched to Elizabeth Hebden -- please verify. "
            "Elizabeth Hebden is also on the Brandon-Elijah FB-invited range.)"
        ),
        message=(
            "It will be a beautiful celebration of you both! "
            "Looking forward to it! Hope my dress still fits! Xx"
        ),
        _src=("form:yes-1-elizabeth.txt", "csv:r6"),
    ),

    # CSV r7-8 -- Claire + her son Elijah (per Anthony 2026-04-07). Both FB invited.
    party(
        "Claire Jopling & Elijah Havard",
        headcount=2,
        source="FB invited (no response)",
        notes="Elijah is Claire's son",
        _src=("csv:r7-8", "fb-invited:Brandon-Elijah"),
    ),

    # CSV r9-10 -- Dave Probert + Wendy share dprobert@pc-q.net
    # Julie's email: "Dave (allow for +1, need to ring and chase him)"
    party(
        "Dave Probert & Wendy",
        headcount=2,
        contact="dprobert@pc-q.net",
        rsvp="awaiting",
        notes=(
            "Need to ring and chase. "
            "(Julie's email referred to 'Dave (+1)' -- assumed = Dave Probert + Wendy from CSV.)"
        ),
        _src=("csv:r9-10", "julie-email:#4"),
    ),

    # CSV r11 -- Paola, also form respondent
    party(
        "Paola Marcon",
        headcount=1,
        contact="paola.marcon@gmail.com",
        rsvp="yes",
        source="form",
        museum="yes",
        _src=("form:yes-1-paola_marcon.txt", "csv:r11"),
    ),

    # CSV r12 -- Annie Levy, FB=Y in CSV but ALSO submitted form -> trust form
    party(
        "Annie Levy",
        headcount=1,
        contact="annielevyevermore@gmail.com",
        rsvp="yes",
        source="form",
        museum="no",
        notes="(also marked accepted via FB in Julie's CSV)",
        _src=("form:yes-1-annie_levy.txt", "csv:r12"),
    ),

    # CSV r13-14 -- Cheryl (Anthony's mother) + Frank, no surnames in CSV
    # Julie's email: "Cheryl and Frank (confirmed - staying at Old Vicarage, Dolfor)"
    party(
        "Cheryl & Frank",
        headcount=2,
        rsvp="yes",
        accommodation="Old Vicarage, Dolfor",
        notes="Anthony's mother and her husband Frank. Confirmed via Julie (channel unspecified).",
        _src=("csv:r13-14", "julie-email:#15"),
    ),

    # CSV r15-16 -- Geoff Bailey + Sopha = "Dad and Sopha" form
    party(
        "Geoff Bailey & Sopha",
        headcount=2,
        contact="geoffxbailey@cloud.com",
        rsvp="yes",
        source="form",
        museum="yes",
        notes="Anthony's father and his wife Sopha",
        _src=("form:yes-2-dad_and_sopha.txt", "csv:r15-16"),
    ),

    # CSV r17-18 -- Stephen Bailey + Fran = Francesca Tambellini
    party(
        "Stephen Bailey & Francesca Tambellini",
        headcount=2,
        contact="stephen.aj.bailey@gmail.com",
        rsvp="yes",
        source="form",
        museum="yes",
        dietary="Vegetarian",
        notes="Anthony's brother and his partner Francesca. CSV note: 'veggie, dog?'",
        message=(
            "Delighted to be coming. Let us know about dog when you can. "
            "It's totally fine to leave him with friends. xx"
        ),
        _src=("form:yes-2-stephen_bailey+francesca_tambellini.txt", "csv:r17-18"),
    ),

    # CSV r19 -- Simon Bailey, declined via WhatsApp, headcount 0
    party(
        "Simon Bailey",
        headcount=0,
        rsvp="no",
        source="text",
        notes=(
            "Invited on WhatsApp, sent apologies, doesn't expect to be able to come. "
            "(per CSV note)"
        ),
        _src=("csv:r19",),
    ),

    # CSV r20-21 -- Ali Morrison = Alexandra Morrison (form)
    party(
        "Alexandra (Ali) & Richard Morrison",
        headcount=2,
        contact="alimorrison70@hotmail.co.uk",
        rsvp="yes",
        source="form",
        museum="yes",
        message=(
            "We would be absolutely delighted to be included in your celebrations. "
            "You can't rush these things but glad you've finally decided to take the plunge xx"
        ),
        _src=("form:yes-2-alexandra_morrison+richard_morrison.txt", "csv:r20-21"),
    ),

    # CSV r22 -- Shazia Akbar = "Shaz" in Julie's email
    party(
        "Shazia (Shaz) Akbar",
        headcount=1,
        rsvp="awaiting",
        dietary="Halal",
        travel="Can drive if necessary",
        notes="(per Julie's email)",
        _src=("csv:r22", "julie-email:#17"),
    ),

    # CSV r23-24 -- John Hale (III) + Antony White, both FB=Y but also form
    party(
        "Antony White & John Hale III",
        headcount=2,
        contact="Antonywhite@me.com",
        rsvp="yes",
        source="form",
        museum="yes",
        notes="(also marked accepted via FB in Julie's CSV)",
        message="We'll be there. So excited to share your big day with you both x",
        _src=("form:yes-2-antony_white+john_hale_iii.txt", "csv:r23-24"),
    ),

    # CSV r25 -- Alley Wender, FB-confirmed, AM=M
    # Julie's email: confirmed, on FB event but not website
    party(
        "Alley Wender",
        headcount=1,
        rsvp="yes",
        source="Facebook",
        museum="maybe",
        notes=(
            "Not good with tech so not RSVPed through website but has confirmed. "
            "(per CSV / Julie's email)"
        ),
        _src=("csv:r25", "julie-email:#1"),
    ),

    # CSV r26 -- Amanda James, ODW=Y FB=Y, form yes-2 (no +1 named)
    party(
        "Amanda James (+1)",
        headcount=2,
        contact="Amandacarms@outlook.com",
        rsvp="yes",
        source="form",
        museum="yes",
        notes="Form indicates party of 2 but +1 not named",
        message="So happy for you xxa",
        _src=("form:yes-2-amanda_james.txt", "csv:r26"),
    ),

    # CSV r27-28 -- Suzanne Marshall Smith + Bryan Marshall (TENTATIVE pair)
    party(
        "Suzanne Marshall Smith & Bryan Marshall",
        headcount=2,
        rsvp="yes",
        source="Facebook",
        notes="(tentatively grouped as a couple from adjacent CSV rows + shared 'Marshall' surname -- please verify)",
        _src=("csv:r27-28",),
    ),

    # CSV r29 -- Andrew Deathe, also form respondent
    party(
        "Andrew Deathe",
        headcount=1,
        contact="07980077179",
        rsvp="yes",
        source="form",
        museum="maybe",
        message="Long life, good health and boundless happiness to you both.",
        _src=("form:yes-1-andrew_deathe.txt", "csv:r29"),
    ),

    # CSV r30 -- Judy Mitchinson, FB=Y
    party(
        "Judy Mitchinson",
        headcount=1,
        rsvp="yes",
        source="Facebook",
        _src=("csv:r30",),
    ),

    # CSV r31 -- KD Shull, FB=Y in CSV, also form yes-3 with kids
    party(
        "KD Shull, Oliver & Leo (kids)",
        headcount=3,
        contact="kd_shull@hotmail.com",
        rsvp="yes",
        source="form",
        museum="yes",
        dietary="No seafood (not an allergy)",
        notes="(also marked accepted via FB in Julie's CSV; CSV doesn't list the kids)",
        _src=("form:yes-3-kd_shull+oliver+leo_kids.txt", "csv:r31"),
    ),

    # CSV r32-33 -- Deborah Simmonds + Tod Simmonds (form: Deb + Tod)
    party(
        "Deborah Simmonds & Tod",
        headcount=2,
        contact="Debssimm@aol.com",
        rsvp="yes",
        source="form",
        museum="yes",
        notes="(Deborah also marked accepted via FB in Julie's CSV)",
        message="See you there!",
        _src=("form:yes-2-deb_simmonds+tod.txt", "csv:r32-33"),
    ),

    # CSV r34-35 -- Wolfgang + Kingsley are a couple (per Anthony 2026-04-07).
    # Both FB=Y in CSV (accepted via Facebook).
    party(
        "Wolfgang Schaeffer & Kingsley George",
        headcount=2,
        rsvp="yes",
        source="Facebook",
        notes="(Kingsley George is NOT the 'George' museum volunteer in Julie's email -- that's a separate row)",
        _src=("csv:r34-35",),
    ),

    # CSV r36 -- Nick Laing, also form
    party(
        "Nick Laing",
        headcount=1,
        contact="Nickdaniellaing@gmail.com",
        rsvp="yes",
        source="form",
        museum="maybe",
        message="Hope you have a beautiful day and look forward to seeing you",
        _src=("form:yes-1-nick_laing.txt", "csv:r36"),
    ),

    # CSV r37 -- Joshua Cobbs, form: no
    party(
        "Joshua Cobbs",
        headcount=0,
        contact="jjcobbs@gmail.com",
        rsvp="no",
        source="form",
        message=(
            "Unfortunately we are unable to help you celebrate. "
            "We wish you all the best and hope our paths will cross sooner rather than later!"
        ),
        _src=("form:no-0-joshua_cobbs.txt", "csv:r37"),
    ),

    # CSV r38 -- FB invited, no response
    party(
        "Benjamin Rinehart",
        source="FB invited (no response)",
        _src=("csv:r38", "fb-invited:Judy-Carmen"),
    ),

    # CSV r39 -- Mark Gowdy, form: no
    party(
        "Mark Gowdy",
        headcount=0,
        contact="markgowdy@gmail.com",
        rsvp="no",
        source="form",
        message=(
            "I'm so sorry but I will not be able to make it over to the UK at that time. "
            "I am sure it is going to be a wonderful day."
        ),
        _src=("form:no-0-mark_gowdy.txt", "csv:r39"),
    ),

    # CSV r40-41 -- Rosie Hammick + Oli Ferguson, also form
    party(
        "Rosie Hammick & Oli Ferguson",
        headcount=2,
        contact="rosiehammick@gmail.com",
        rsvp="yes",
        source="form",
        museum="yes",
        dietary=(
            "Rosie intolerances - garlic, fructose, beans/lentils. "
            "Although please don't worry too much about me -- it's a tricky one. "
            "Hoping to sort with antibiotics in the next few months anyway."
        ),
        message=(
            "Dear Julie & Anthony, Oli and I would like to celebrate with you on your "
            "special wedding day - thank you for the invite. We hope you're both well "
            "and look forward to hopefully seeing you soon. Rosie & Oli xx"
        ),
        _src=("form:yes-2-rosie_hammick+oli_ferguson.txt", "csv:r40-41"),
    ),

    # CSV r42 -- Katharine Brown, form
    party(
        "Katharine Brown",
        headcount=1,
        contact="07796921502",
        rsvp="yes",
        source="form",
        museum="yes",
        dietary="Vegetarian",
        notes="Chris (partner) can't make it -- putting on a festival that day",
        message=(
            "Congratulations! Thanks so much for inviting me. "
            "Chris can't make it as he is putting on a festival that day, "
            "but I would love to join you. X"
        ),
        _src=("form:yes-1-katharine_brown.txt", "csv:r42"),
    ),

    # CSV r43 -- Jamie Burt, form
    party(
        "Jamie Burt",
        headcount=1,
        contact="Jamieburt25@gmail.com",
        rsvp="yes",
        source="form",
        museum="yes",
        message="Yay! Looking forward to celebrating love with you both.",
        _src=("form:yes-1-jamie_burt.txt", "csv:r43"),
    ),

    # CSV r44 -- FB invited, no response
    party(
        "Sienna Holmes",
        source="FB invited (no response)",
        _src=("csv:r44", "fb-invited:Judy-Carmen"),
    ),

    # CSV r45 -- FB invited, no response
    party(
        "Annie Kilby",
        source="FB invited (no response)",
        _src=("csv:r45", "fb-invited:Judy-Carmen"),
    ),

    # CSV r46 -- Jane Rigby, form: maybe
    party(
        "Jane Rigby",
        headcount=1,
        contact="07771 590883 / janeybean65@gmail.com",
        rsvp="maybe",
        source="form",
        museum="maybe",
        notes="Probably away in Outer Hebrides 15-27 May -- 'leaving a corner of mind open'",
        message=(
            "Wow, I am so happy for you both. It's the most beautiful wedding scenario "
            "of you two, and you have created and chosen the most incredible celebrations "
            "and such a joyous feel. I wish I could come, but I'm expecting to be away "
            "from the 15th to 27th May in the Outer Hebrides and it's too far to go "
            "to come back after a week. BUT I am learning that the best laid plans don't "
            "always turn out the way I expect. So I am wanting to leave a little corner "
            "of my mind open to the possibility of being able to come? But I will be "
            "most likely unable to attend, I'm so sorry. And thank you so much for asking "
            "me, I'm very chuffed that you did. Much love to you both and I will be "
            "making a something for you! XXXXX"
        ),
        _src=("form:maybe-1-jane_rigby.txt", "csv:r46"),
    ),

    # CSV r47 -- Jane Baker, form
    party(
        "Jane Baker",
        headcount=1,
        contact="janebaker81@googlemail.com",
        rsvp="yes",
        source="form",
        museum="yes",
        message="Whoop whoop! Big congratulations xx",
        _src=("form:yes-1-jane_baker.txt", "csv:r47"),
    ),

    # CSV r48 -- "Debbie" -- possibly = Deborah Simmonds duplicate? FB invited.
    party(
        "Debbie",
        source="FB invited (no response)",
        notes="(possibly duplicate of Deborah Simmonds row -- please verify)",
        _src=("csv:r48", "fb-invited:Judy-Carmen"),
    ),

    # CSV r49 -- Carmen on FB invited list. Julie's email adds Mike.
    party(
        "Mike & Carmen",
        headcount=2,
        rsvp="awaiting",
        source="FB invited (no response)",
        notes=(
            "Carmen is the FB invitee. Mike is not in the CSV, "
            "added per Julie's email '#14 Mike and Carmen'."
        ),
        _src=("csv:r49", "julie-email:#14", "fb-invited:Judy-Carmen"),
    ),

    # CSV r50 -- Sheila (no surname); Julie's email = Sheila Timms
    party(
        "Sheila Timms",
        headcount=2,
        rsvp="awaiting",
        notes=(
            "Need to email for RSVP. +1 allowed in case she brings a driver. "
            "(Julie's email; CSV row contact is just 'text')"
        ),
        _src=("csv:r50", "julie-email:#2"),
    ),

    # CSV r51 -- Kerry Henderson, with email contact
    party(
        "Kerry Henderson",
        headcount=1,
        contact="kerrylou@hotmail.co.uk",
        rsvp="awaiting",
        source="email",
        notes="Emailed, no response so probably no. (per Julie's email)",
        _src=("csv:r51", "julie-email:#5"),
    ),

    # CSV r52-53 -- Andrew Logan + Michael Davies = "Andrew and Michael" in Julie's email
    # (Andrew Logan is the sculptor, the museum venue's namesake!)
    party(
        "Andrew Logan & Michael Davies",
        headcount=2,
        contact="michaelsjd@mac.com",
        rsvp="awaiting",
        notes="Andrew Logan = the museum's namesake sculptor",
        _src=("csv:r52-53", "julie-email:#3"),
    ),

    # CSV r54-55 -- Ash + "Amber" but per Anthony the couple is Ash + Alexis
    party(
        "Ash & Alexis",
        headcount=2,
        rsvp="awaiting",
        travel="Driving from Montgomery",
        notes=(
            "CSV row 55 says 'Amber' but Anthony confirms the couple is "
            "Ash + Alexis -- please verify with Julie."
        ),
        _src=("csv:r54-55", "julie-email:#16"),
    ),

    # CSV r56 -- Alisdair Tullo, form
    party(
        "Alisdair Tullo",
        headcount=1,
        contact="alisdair@tullo.me.uk",
        rsvp="yes",
        source="form",
        museum="yes",
        dietary=(
            "Gluten free (the family curse I've been ignoring up til now). "
            "Weirdly allergic to buckwheat too (obscure I know)"
        ),
        message="Wouldn't miss it! Love to you both.",
        _src=("form:yes-1-alisdair_tullo.txt", "csv:r56"),
    ),

    # ========================================================================
    # CSV r57-61 -- RESERVE LIST (per Anthony 2026-04-07)
    # ========================================================================
    party("Ralph", notes="Reserve list", _src=("csv:r57",)),
    party("Rhys", notes="Reserve list", _src=("csv:r58",)),
    party("Diane", notes="Reserve list", _src=("csv:r59",)),
    party("Tammi Dallaston", notes="Reserve list", _src=("csv:r60",)),
    party("Terri Wilson", notes="Reserve list", _src=("csv:r61",)),

    # ========================================================================
    # CSV later rows -- Emma Fork section (NOT reserve list per Anthony)
    # ========================================================================

    # Emma Fork has FB=M (maybe via Facebook)
    party(
        "Emma Fork",
        headcount=1,
        rsvp="maybe",
        source="Facebook",
        _src=("csv:Emma Fork",),
    ),

    party(
        "Susan Hutchison",
        source="FB invited (no response)",
        _src=("csv:Susan Hutchison", "fb-invited:Emma-Jane"),
    ),
    party(
        "Jane Reynolds Culliford",
        source="FB invited (no response)",
        _src=("csv:Jane Reynolds Culliford", "fb-invited:Emma-Jane"),
    ),

    # ========================================================================
    # Julie's email parties NOT in CSV (#6, 7, 8, 9, 10, 11, 12, 13)
    # ========================================================================
    party(
        "Sharon",
        headcount=1,
        rsvp="yes",
        notes="Volunteer from Andrew Logan Museum (per Julie's email)",
        _src=("julie-email:#6",),
    ),
    party(
        "George",
        headcount=1,
        rsvp="yes",
        notes=(
            "Volunteer from Andrew Logan Museum (per Julie's email). "
            "NOT the same person as Kingsley George above."
        ),
        _src=("julie-email:#7",),
    ),
    party(
        "Pip",
        headcount=1,
        rsvp="yes",
        notes="'coordinating? to meet at weekend - confirmed' (per Julie's email)",
        _src=("julie-email:#8",),
    ),
    party(
        "Simon Campion (+1)",
        headcount=2,
        rsvp="awaiting",
        notes="+1 allowed (per Julie's email)",
        _src=("julie-email:#9",),
    ),
    party("Sean Young", rsvp="awaiting", _src=("julie-email:#10",)),
    party("Janet Slee", rsvp="awaiting", _src=("julie-email:#11",)),
    party("Dominic Ashmole", rsvp="awaiting", _src=("julie-email:#12",)),
    party(
        "Raif (+4)",
        headcount=5,
        rsvp="awaiting",
        notes="Raif +4 (per Julie's email)",
        _src=("julie-email:#13",),
    ),

    # ========================================================================
    # Form-only test entry (Claude) -- kept per Anthony 2026-04-07
    # ========================================================================
    party(
        "Claude (test)",
        headcount=1,
        contact="noreply@anthropic.com",
        rsvp="maybe",
        source="form",
        museum="yes",
        notes="Form test entry from Feb 2026",
        message=(
            "I would love to be there but I am not entirely sure I will have a body "
            "by May 2026. Congratulations to you both!"
        ),
        _src=("form:maybe-1-claude.txt",),
    ),
]


# ----------------------------------------------------------------------------
# XLSX GENERATION
# ----------------------------------------------------------------------------

def generate():
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()

    # Sheet 1: Guest list ----------------------------------------------------
    ws = wb.active
    ws.title = "Guest list"

    # Header row
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="DDEEDD", end_color="DDEEDD", fill_type="solid")
    for col_idx, name in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.freeze_panes = "A2"

    # Data rows
    field_order = [
        "party", "headcount", "contact", "rsvp", "source",
        "museum", "dietary", "accommodation", "travel", "notes", "message",
    ]
    for row_idx, p in enumerate(PARTIES, start=2):
        for col_idx, field in enumerate(field_order, start=1):
            ws.cell(row=row_idx, column=col_idx, value=p[field])

    # Column widths (rough)
    widths = {
        "Party": 32, "Headcount": 10, "Contact": 30, "RSVP": 10,
        "Source": 12, "Museum (AM)": 12, "Dietary": 30,
        "Accommodation": 22, "Travel": 22, "Notes": 40, "Message": 50,
    }
    for col_idx, name in enumerate(COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = widths[name]

    # Data validation dropdowns (apply to a generous row range to cover
    # future additions in the sheet)
    last_row = max(500, len(PARTIES) + 100)

    def add_dropdown(col_letter, values):
        formula = '"' + ",".join(values) + '"'
        dv = DataValidation(type="list", formula1=formula, allow_blank=True)
        dv.error = "Pick from the dropdown"
        dv.errorTitle = "Invalid value"
        ws.add_data_validation(dv)
        dv.add(f"{col_letter}2:{col_letter}{last_row}")

    add_dropdown("D", RSVP_VALUES)        # RSVP
    add_dropdown("E", SOURCE_VALUES)      # Source
    add_dropdown("F", MUSEUM_VALUES)      # Museum (AM)

    # Sheet 2: Links ---------------------------------------------------------
    ws2 = wb.create_sheet("Links")
    ws2.cell(row=1, column=1, value="Resource").font = header_font
    ws2.cell(row=1, column=2, value="URL").font = header_font
    ws2.cell(row=1, column=1).fill = header_fill
    ws2.cell(row=1, column=2).fill = header_fill
    for i, (label, url) in enumerate(LINKS, start=2):
        ws2.cell(row=i, column=1, value=label)
        ws2.cell(row=i, column=2, value=url)
    ws2.column_dimensions["A"].width = 24
    ws2.column_dimensions["B"].width = 60

    wb.save(OUTPUT)
    print(f"Wrote {OUTPUT} ({len(PARTIES)} parties)")


if __name__ == "__main__":
    generate()
